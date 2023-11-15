import matplotlib.pyplot as plt
import openpyxl
import os
from matplotlib.font_manager import FontProperties

def add_x_line(plt, desired_x):
    """
    Adds a vertical dashed line to a given Matplotlib plot.

    Parameters:
    - plt (Matplotlib plot): The plot to which the line will be added.
    - desired_x (float): The x-value at which the vertical dashed line should be drawn.

    Returns:
    - plt (Matplotlib plot): The updated plot with the added line.
    """
    plt.axvline(x=desired_x, color='gray', linestyle='--')
    return plt

def find_closest_value(lst, target):
    """
    Find the value in a list that is closest to a given integer.

    Parameters:
    - lst (list): A list of numerical values.
    - target (int): The integer to which you want to find the closest value.

    Returns:
    - int or float: The value from the list that is closest to the target integer.
    """
    return min(lst, key=lambda x: abs(x - target))

def find_first_numeric_row(ws, column_index):
    for row_index, cell in enumerate(ws.iter_rows(min_col=column_index, max_col=column_index, values_only=True), start=1):
        if cell[0] is not None and isinstance(cell[0], (int, float)):
            return row_index

    # Return a default value or raise an exception if no numeric value is found
    return None

def plot_excel_data(plt, 
                    excel_path,
                    image_size,
                    colorMap = "",
                    xes_to_highlight = [], 
                    draw_highlight_line = True, 
                    use_grid = True, 
                    savefig = False, 
                    axis_label_font_size = 12, 
                    text_font_size = 10):
    """
    Plots data from an Excel file using Matplotlib.

    Parameters:
    - plt (Matplotlib plot): The plot to be used for displaying the data.
    - excel_path (str): The path to the Excel file or directory containing Excel files to be plotted.
    - x_to_highlight (float): The x-value at which to highlight data points.
    - image_size (list): The size of the plot image as [width, height].
    - draw_highlight_line (bool): Whether to draw a vertical dashed line at the x_to_highlight value.
    - use_grid (bool): Whether to display a grid on the plot.
    - savefig (bool): Whether to save the plot as an image file.
    - axis_label_font (tuple): Font information for axis labels as (fontname, fontsize).
    - text_font (tuple): Font information for text on the plot as (fontname, fontsize).

    Functionality:
    - Reads data from the specified Excel file(s).
    - Plots the data using Matplotlib, with options for customization.
    - Adds a vertical dashed line at the specified x-value if draw_highlight_line is True.
    - Displays a grid if use_grid is True.
    - Saves the plot as an image file if savefig is True.
    """

    excel_list = []
    excel_filename_list = []

    if ".xlsx" in excel_path:
        excel_list.append(excel_path)
        excel_filename_list.append(excel_path.split("/")[-1].split("\\")[-1])
    else:
        # Iterate over all files in the directory
        for root, dirs, files in os.walk(excel_path):
            for file in files:
                if file.endswith(".xlsx"):
                    excel_list.append(os.path.join(root, file))
                    excel_filename_list.append(file)

    for idx, excel_file in enumerate(excel_list):
        # Specify the name of the sheet where the data is located
        sheet_name = 'Plot Data'

        filename = excel_filename_list[idx]

        # Open the Excel file using openpyxl
        wb = openpyxl.load_workbook(excel_file)
        ws = wb[sheet_name]

        # Find the columns that contain the string "Physical time [s]" in row 4 and use the text to the right as labels
        time_columns = {}
        for col_index, cell in enumerate(ws[4], 1):
            if cell.value == "Physical time [s]":
                time_columns[col_index] = ws.cell(row=4, column=col_index + 1).value

        # Extract the data from the selected time columns and their corresponding data columns
        column_data = []
        column_times = []
        column_labels = []

        for time_column, label in time_columns.items():
            
            # Find the first row with data
            first_row = find_first_numeric_row(ws, time_column)

            column_times.append([cell[0].value for cell in ws.iter_rows(min_row=first_row, max_row=ws.max_row, min_col=time_column, max_col=time_column)])
            data_column = time_column + 1
            column_data.append([cell[0].value for cell in ws.iter_rows(min_row=first_row, max_row=ws.max_row, min_col=data_column, max_col=data_column)])
            column_labels.append(label)

            print(first_row)

        # Create a line graph using matplotlib with custom figsize and tight layout.
        plt.figure(figsize=(image_size[0], image_size[1]), tight_layout=True)
        
        # Set the default color cycle if the user have given us an input
        if not (colorMap == ""):
            plt.rcParams['axes.prop_cycle'] = plt.cycler(color=plt.cm.get_cmap(colorMap).colors)


        # Set font path
        font_path = r'C:\Users\avalonuser\Downloads\Montserrat\static\Montserrat-Regular.ttf'
        font = FontProperties(fname=font_path)

        max_len = 0

        # Plot the data from the selected columns and use the labels from the right as the legend
        for i in range(len(column_data)):

            x = [z for z in column_times[i] if z is not None]
            y = [z for z in column_data[i] if z is not None]

            plt.plot(x, y, label=column_labels[i])

            last_value = y[-1]
            color = plt.gca().get_lines()[-1].get_color()
            plt.text(x[-1], last_value, str(int(last_value)), color=color, va='bottom', fontproperties=font, fontsize=text_font_size)

            # Get the longest series for plotting highlight y values
            if max(x) > max_len:
                max_len = max(x)


        # Set the chart title and axis labels
        plt.title(filename)

        plt.xlabel('Time [s]', fontproperties=font, fontsize=axis_label_font_size)
        plt.ylabel('Temperature [°C]', fontproperties=font, fontsize=axis_label_font_size)

        # Show the graph with the legend    
        legend = plt.legend()
        plt.legend(loc='upper left') # Maybe want it somewhere else? However, it may overlap with the averages
        legend.get_frame().set_alpha(0.5)  # You can adjust the alpha (0.0 to 1.0) to control the level of transparency

        # Loop over all given x-values to plot
        for x_to_highlight in xes_to_highlight:
            text_count = 0

            # Find the closest x-value for the given x.
            x_to_highlight = find_closest_value(x, x_to_highlight)

            # Print the y-values for all lines at the desired x-value
            for i, x_value in enumerate(x):  # Using the first column for x-values
                if x_value == x_to_highlight:
                    for j, column_label in enumerate(column_labels):
                        y_value = column_data[j][i]

                        text_y = 15 + (text_count * 8)  # Adjust vertical offset

                        color = plt.gca().get_lines()[j].get_color()

                        # Place the text in relation to the line, and the length of our series.
                        plt.text(x_to_highlight + (max_len*0.005), text_y, f'{y_value:.3g}', color=color, va='bottom', fontproperties=font, fontsize=text_font_size)
                        text_count += 1

                        print(f'x={x_to_highlight:.3g}, y={y_value:.3g}')

            # Only draw the line if the user wants it
            if draw_highlight_line:
                plt = add_x_line(plt, x_to_highlight)

        if use_grid:
            plt.grid()

        if savefig:
            path = excel_file.split(".xlsx")[0]
            plt.savefig(f'{path}.png', transparent=True)
            print(f'Saved file: {path}.png')

        plt.show()

        print("---------------------")


plot_excel_data(plt,
                r'C:\Users\avalonuser\Downloads\20 kW olika centrum och rh (1).xlsx',
                colorMap = "Set2",
                image_size = [12, 6],
                xes_to_highlight = [6.2],
                draw_highlight_line = True,
                use_grid = True,
                savefig = True,
                axis_label_font_size = 12,
                text_font_size = 12)