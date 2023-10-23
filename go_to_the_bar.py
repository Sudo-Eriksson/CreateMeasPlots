import openpyxl
import matplotlib.pyplot as plt
import numpy as np
import seaborn as sns
import pandas as pd


def find_start_row(sheet):
    """
    Find the first row in a given Excel sheet containing numeric data in the second column.

    Parameters:
    - sheet (openpyxl.worksheet.worksheet.Worksheet): The Excel sheet to search for numeric data.

    Returns:
    - int or None: The row number of the first row with numeric data in the second column, or None if no such row is found.

    This function searches for the first row in the provided Excel sheet where the second column contains numeric data (integer or float). It is particularly useful for locating the starting point for data extraction when the sheet may contain non-numeric header rows.
    If a row with numeric data in the second column is found, the function returns the row number. If no numeric data is found, it returns None.
    """
    # Loop through rows in the Excel sheet and find the first row with a number in column 2
    for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
        if isinstance(row[1], (int, float)):
            return row_num
    return None

def create_bar_chart(file_path, figure_size=(10, 6), savefig=False, text_size=12, text_font='sans-serif', create_score=False):
    """
    Create a bar chart from data in an Excel file.

    Parameters:
    - file_path (str): The path to the Excel file containing the data.
    - figure_size (tuple): Optional. A tuple specifying the size of the figure (width, height) in inches.
    - savefig (bool): Whether to save the plot as an image file.
    - text_size (int): Optional. The font size for the text labels above the bars.
    - text_font (str): Optional. The font style for the text labels above the bars.
    - create_score (bool): Whether to subplot a evenness and level score.

    This function reads data from the Excel file and creates a bar chart with three bars (Min, Mean, Max) per data point.
    The chart includes horizontal gridlines and labels. You can customize the figure size using the figure_size parameter.
    """
    # Open the Excel file
    excel_file = openpyxl.load_workbook(file_path)

    for sheet in excel_file.sheetnames:
        current_sheet = excel_file[sheet]
        start_row = find_start_row(current_sheet)
        if start_row is None:
            print("No numeric data found in column 2.")
            return

        names = []
        min_value = []
        mean_value = []
        max_value = []

        for row in current_sheet.iter_rows(min_row=2, values_only=True):
            names.append(row[0])
            min_value.append(row[1])
            mean_value.append(row[2])
            max_value.append(row[3])

        num_names = len(names)
        width = 0.175
        spacing = 0.2
        x = np.arange(num_names)

        plt.figure(figsize=figure_size)

        # Only create subplot if this bool is true
        if create_score:
            ax1 = plt.subplot(211)  # Skapa den första subplotten (2 rader, 1 kolumn, första plats)
        
        plt.bar(x - spacing, min_value, width=width, label='Min', color='darkslategrey')
        plt.bar(x, mean_value, width=width, label='Mean', color='goldenrod')
        plt.bar(x + spacing, max_value, width=width, label='Max', color='firebrick')

        plt.grid(axis='y', linestyle='--', alpha=0.7)
        plt.xticks(x, names)
        filename = file_path.split("/")[-1].split("\\")[-1]
        plt.ylabel('Temperature [°C]')
        plt.title(f'{filename}: {sheet}')
        plt.legend()
        delta = 0.00

        for i in range(num_names):
            plt.text(x[i] - spacing - delta, min_value[i], f"{min_value[i]:.3g}", ha='center', va='bottom',
                    color='darkslategrey', fontsize=text_size, family=text_font)
            plt.text(x[i] - delta, mean_value[i], f"{mean_value[i]:.3g}", ha='center', va='bottom',
                    color='goldenrod', fontsize=text_size, family=text_font)
            plt.text(x[i] + spacing - delta, max_value[i], f"{max_value[i]:.3g}", ha='center', va='bottom',
                    color='firebrick', fontsize=text_size, family=text_font)

        # Create score subplot if wanted by user.
        if create_score:
            # Lägg till en subplott för (max - min) / mean under stapeldiagrammet
            ax2 = plt.subplot(212, sharex=ax1)  # Skapa den andra subplotten (2 rader, 1 kolumn, andra plats)
            #ratio = [(max_val - min_val) / mean_val for max_val, min_val, mean_val in zip(max_value, min_value, mean_value)]
            ratio = [(((max_val - min_val) + (max_val - mean_val) + (mean_val - min_val) )/ 3)/(mean_val) for max_val, min_val, mean_val in zip(max_value, min_value, mean_value)]
            plt.bar(x, ratio, width=width, label='Evenness and level score', color='royalblue')
            plt.ylabel('Evenness and level score')
            plt.legend()
            plt.grid(axis='y', linestyle='--', alpha=0.7)

            for i in range(num_names):
                plt.text(x[i], ratio[i], f"{ratio[i]:.3g}",
                        ha='center', va='bottom', color='royalblue', fontsize=text_size, family=text_font)

        if savefig:
            path = file_path.split(".xlsx")[0]
            plt.savefig(f'{path}_{sheet}.png', transparent=True)
            print(f'Saved file: {path}_{sheet}.png')

        plt.show()

def create_radar_subplots(file_path):
    """
    create_radar_subplots(file_path)

    Description:
    This function creates radar chart subplots from data stored in an Excel file. It extracts information about different entities (e.g., products, categories) and their corresponding minimum, mean, and maximum values. Each entity's data is plotted as a radar chart in a subplot. Subplots are organized in rows with a specified number of subplots per row.

    Parameters:
    - file_path (str): The file path to the Excel file containing the data to be visualized.

    Returns:
    None
    """

    # Open the Excel file
    excel_file = openpyxl.load_workbook(file_path)
    sheet = excel_file.active

    # Create empty lists for data
    names = []
    min_value = []
    mean_value = []
    max_value = []

    # Loop through rows in the Excel sheet and retrieve data
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2 to skip the headers
        names.append(row[0])
        min_value.append(row[1])
        mean_value.append(row[2])
        max_value.append(row[3])

    num_datapoints = len(names)

    # Set the number of subplots per row
    subplots_per_row = 4

    # Calculate the number of rows needed
    num_rows = (num_datapoints + subplots_per_row - 1) // subplots_per_row

    # Create subplots for radar charts
    fig, axs = plt.subplots(num_rows, subplots_per_row, subplot_kw=dict(polar=True), figsize=(15, 5 * num_rows))

    # Adjust the spacing between subplots
    plt.subplots_adjust(hspace=0.5)

    # Determine the maximum value for the radar plot
    max_radar_value = max(max_value)

    for i in range(num_datapoints):
        # Create a list of labels for each category
        categories = ['Min', 'Mean', 'Max']

        # Create values for the radar chart
        values = [min_value[i], mean_value[i], max_value[i]]

        # Duplicate the first value to close the circular graph
        values += values[:1]

        # Calculate angles for each category
        angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()

        # Add the first value at the end to complete the circular plot
        angles += angles[:1]

        # Create the radar chart
        row = i // subplots_per_row
        col = i % subplots_per_row
        ax = axs[row, col]
        ax.fill(angles, values, 'b', alpha=0.1)
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(categories)
        ax.set_yticklabels([])
        ax.set_title(names[i])

        # Set the same maximum value for all radar plots
        ax.set_rmax(max_radar_value)

    # Remove any empty subplots
    for i in range(num_datapoints, num_rows * subplots_per_row):
        row = i // subplots_per_row
        col = i % subplots_per_row
        fig.delaxes(axs[row, col])

    # Display the radar subplots
    plt.show()

def create_seaborn_combined_bar_chart(file_path, figure_size=(10, 6), savefig=False, text_size=12, text_font='sans-serif'):
    # Open the Excel file
    excel_file = openpyxl.load_workbook(file_path)

    for sheet in excel_file.sheetnames:
        current_sheet = excel_file[sheet]
        start_row = find_start_row(current_sheet)
        if start_row is None:
            print("No numeric data found in column 2.")
            return

        names = []
        min_value = []
        mean_value = []
        max_value = []

        for row in current_sheet.iter_rows(min_row=2, values_only=True):
            names.append(row[0])
            min_value.append(row[1])
            mean_value.append(row[2])
            max_value.append(row[3])

        data = pd.DataFrame({'Names': names, 'Min': min_value, 'Mean': mean_value, 'Max': max_value})

        plt.figure(figsize=figure_size, tight_layout=True)

        sns.set_style("whitegrid")
        ax = sns.barplot(x='Names', y='Mean', data=data, color='royalblue', alpha=0.7, label='Mean')
        sns.despine(left=True)

        # Plot vertical lines for Min and Max
        for i in range(len(names)):
            plt.vlines(x=i, ymin=min_value[i], ymax=max_value[i], color='black', linewidth=2)

            # Print mean value above the bar
            ax.text(i + 0.2, mean_value[i], f"{mean_value[i]:.3g}", ha='center', va='bottom', color='royalblue',
                    fontsize=text_size, family=text_font)

            # Print min value above the vertical line
            ax.text(i, max_value[i] + 0.5, f"{max_value[i]:.3g}", ha='center', va='bottom', color='black',
                    fontsize=text_size, family=text_font)

            # Print max value below the vertical line
            ax.text(i, min_value[i] - 0.5, f"{min_value[i]:.3g}", ha='center', va='top', color='black',
                    fontsize=text_size, family=text_font)
        
        ax.set_xlabel('')  # Remove x-axis label
        plt.xticks(rotation=45, ha='right')
        filename = file_path.split("/")[-1].split("\\")[-1]
        plt.ylabel('Temperature [°C]')
        plt.title(f'{filename}: {sheet}')
        if savefig:
            path = file_path.split(".xlsx")[0]
            plt.savefig(f'{path}_{sheet}.png', transparent=True)
            print(f'Saved file: {path}_{sheet}.png')

        plt.show()


# Example usage
create_seaborn_combined_bar_chart(r'C:\Users\avalonuser\Desktop\Ytter- och centrumtemp.xlsx',
                                  figure_size = (8, 7),
                                  savefig = True,
                                  text_size=10, 
                                  text_font='serif')

# Example usage with a custom figure size (e.g., 12x8 inches)
#create_bar_chart(r'C:\Users\avalonuser\Desktop\Ytter- och centrumtemp.xlsx',
#                 figure_size = (16, 8),
#                 savefig = True,
#                 text_size=8, 
#                 text_font='serif',
#                 create_score = False)